from datetime import datetime, timedelta
import selenium
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from threading import Thread
from bs4 import BeautifulSoup
import speech_recognition as sr
import lxml
from openpyxl import Workbook, load_workbook
import pyaudio
import speech_recognition as sr


class Schedule:

    def __init__(self, id, code, entry_time, exit_time, driver):
        print(f">new_class created:[{datetime.now()}] code :: {code} :: entry_time :: {entry_time} :: exit_time :: {exit_time} ")
        self.id = id
        self.code = code
        self.entry_time = entry_time
        self.exit_time = exit_time
        self.link = self.links[code]
        self.driver = driver
        self.exited = False
        if self.link is None:
            self.link = input(f"Please enter the link for {self.code} :: ")
        self.link = f"'{self.link}'"

        self.T1 = Thread(target=self.exit_meeting)
        self.T2 = Thread(target=self.attendance)
        self.T3 = Thread(target=self.speech_recognition)
        self.join_meet()

    links = {
        "AP": None,
        "EE": "https://meet.google.com/dju-rikc-dqp",
        "MA": "https://meet.google.com/lookup/egu4culb4h?authuser=0&hs=179",
        "CO": "https://meet.google.com/lookup/eflwgh5lju?authuser=0&hs=179",
        "EELAB": None,
        "COLAB": "https://meet.google.com/lookup/fpslolbjbv?authuser=0&hs=179",
        "APLAB": "https://meet.google.com/fha-actr-rqk",
        "MELAB": "https://meet.google.com/lookup/ay6bqb34ol?authuser=0&hs=179",
        "BCOM": 'https://meet.google.com/tom-cwzr-tce',
        "TESTING": 'https://meet.google.com/qig-heue-mkx'
    }

    xpath = {
        "show_people": '//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[3]/div[2]/div/div/div[2]/span/button/i[1]',
        "mic_main": '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[1]/div[1]/div/div[4]/div[1]/div/div/div/span/span',
        "camera_main": '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[1]/div[1]/div/div[4]/div[2]/div/div/span/span',
        # Xpath for changing background option, only available in the lobby
        "lobby_check": '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[1]/div[1]/div[1]/div[6]/div/span/span/span',
        "join_button": '//*[@id="yDmH0d"]/c-wiz/div/div/div[9]/div[3]/div/div/div[4]/div/div/div[2]/div/div[2]/div/div[1]/div[1]'

    }

    def join_meet(self):
        self.exited = False
        driver = self.driver
        driver.switch_to.window(driver.window_handles[0])
        wait = WebDriverWait(driver, 15)
        driver.execute_script(f"window.open({self.link});")
        tabs = driver.window_handles
        current_tab = tabs[len(tabs) - 1]
        driver.switch_to.window(current_tab)
        mic_main = wait.until(ec.visibility_of_element_located((By.XPATH, self.xpath["mic_main"])))
        action_1 = ActionChains(driver)
        action_1.move_to_element(mic_main).click().perform()
        camera_main = wait.until(ec.visibility_of_element_located((By.XPATH, self.xpath["camera_main"])))
        action_2 = ActionChains(driver)
        action_2.move_to_element(camera_main).click().perform()
        while True:
            try:
                join_button = wait.until(ec.visibility_of_element_located((By.XPATH, self.xpath["join_button"]))).click()
                break
            except selenium.common.exceptions.ElementClickInterceptedException:
                pass
        print(">Waiting in the lobby")
        wait_2 = WebDriverWait(driver=driver, timeout=100000)
        try:
            wait_2.until_not(ec.visibility_of_element_located((By.XPATH, self.xpath["lobby_check"])))
        except selenium.common.exceptions.TimeoutException:
            print(">Class entry denied")
            self.join_meet()
        print(">Class Joined")

        self.T1.start()
        sleep(5)
        self.T2.start()
        self.T3.start()


    def roll_number(self, encore):
        encore = encore.lower()
        encore = encore.split(" ")
        temp = ''
        for split in encore:
            if split.isalpha() is False:
                temp += split

        cross = ''
        for i in temp:
            if i.isnumeric():
                cross += i
            elif i == 'o':
                cross += i

        cross = cross.replace("17", '', 1)
        cross = cross.replace("22", '', 1)
        cross = cross.replace('o', '0')
        if len(cross) == 0:
            cross = '0'

        cross = int(cross)
        if cross == 7500:
            cross = 75
        return cross

    def attendance(self):
        run_interval = 2
        driver = self.driver
        wait = WebDriverWait(driver, 12000000)
        attendance_file = f"{self.code}.xlsx"
        wb = load_workbook(attendance_file)
        ws = wb.active

        # Creating a list of members which already exist in the xl file
        attendee_names_xl = []
        attendees = ws.iter_cols(min_row=2, max_row=ws.find_eod_cell, min_col=1, max_col=1)
        # Even though a singular tuple is being returned in this case
        # The generator objects acts ina  way that multiple tuples are
        # being returned hence we must unpack the first tuple and then access it
        for tuple_1 in attendees:
            attendees = tuple_1

        # Checking for previously existing names in xl file
        for attendee in attendees:
            attendee_names_xl.append(attendee.value)

        # Finding the date column/ Creating one if it doesnt exist already
        max_col = ws.max_column
        dates = ws.iter_rows(min_row=1, max_row=1, min_col=2, max_col=max_col)
        for tuple_1 in dates:
            dates = tuple_1
        date_cell = None

        for date in dates:
            if date.value == self.entry_time:
                date_cell = date
                break

        if date_cell is None:
            date_cell = ws.cell(row=1, column=max_col+1, value=self.entry_time)
        main_col = date_cell.column

        print("")
        print(">", date_cell)
        print(">", date_cell.value)
        print(">Opening attendee list")
        join_button = wait.until(ec.visibility_of_element_located((By.XPATH, self.xpath["show_people"])))
        action_3 = ActionChains(driver)
        action_3.move_to_element(join_button).click().perform()
        sleep(5)

        while self.exited is False:
            start_check = datetime.now()
            print(f">[{datetime.today()}] Checking for attendees")
            names = []
            html_text = driver.page_source
            soup = BeautifulSoup(html_text, 'lxml')
            list_ = soup.find("div", {"role": "list"})
            items = list_.find_all("div", {"role": "listitem"})
            for item in items:
                text = item.text
                names.append(text)

            # adding new attendees to the xl file
            for name in names:
                if name not in attendee_names_xl:
                    new_attendee = ws.cell(row=ws.find_eod_cell + 1, column=1, value=name)
                    attendee_names_xl.append(name)

            # Refreshing the tuple containing cell objects after adding new cells in xl file
            attendees = ws.iter_cols(min_row=2, max_row=ws.find_eod_cell, min_col=1, max_col=1)
            for tuple_1 in attendees:
                attendees = tuple_1

            # Adding time for each attendee
            for attendee in attendees:
                if attendee.value in names:

                    time_add = ws.cell(row=attendee.row, column=main_col)
                    time_elapsed = datetime.now() - start_check
                    time_elapsed = time_elapsed.total_seconds()
                    time_elapsed = time_elapsed/60

                    if time_add.value is None:
                        time_add.value = run_interval + time_elapsed
                    else:
                        time_add.value += run_interval + time_elapsed
                else:
                    pass

            # So that the thread terminates once the class has ended even if its is sleeping for 'run_interval'
            now = datetime.today()
            while datetime.today() - now < timedelta(minutes=run_interval):
                if self.exited is True:
                    break
                else:
                    pass

        # Assigning roll numbers to meet members
        for cell in range(2, ws.find_eod_cell + 1):
            encore = ws['A' + str(cell)].value
            ws['B' + str(cell)].value = self.roll_number(encore)
            print(ws['B' + str(cell)].value)

        wb.save(f"{self.code}.xlsx")
        wb.close()

    def exit_meeting(self):
        now = datetime.today()
        until_exit = self.exit_time - now + timedelta(minutes=3)
        print(f">[{now.time()}] Auto-exit in {until_exit}")

        while datetime.today() - now < until_exit:
            if self.exited is True:
                break
            else:
                pass

        self.exited = True
        self.T2.join()
        self.T3.join()
        self.driver.close()
        print(">Exited class")

    def speech_recognition(self):
        print(">Speech Recognition started")
        recog = sr.Recognizer()
        flag = 1

        while self.exited is False:
            with sr.Microphone() as source:
                audio = recog.listen(source)
                try:
                    text = recog.recognize_google(audio, language='en-IN')
                    print(f"[speech]  {text}")
                except:
                    print("[speech] Failed to recognize")

    def __del__(self):
        if self.T1.is_alive():
            self.T1.join()
        else:
            pass
        print(f">Class object killed: [{datetime.now()}] id :: {self.code} :: entry_time :: {self.entry_time} :: exit_time :: {self.exit_time}")


def open_chrome(profile):
    options = webdriver.ChromeOptions()
    options.add_argument("--user-data-dir=C:/Users/admin/AppData/Local/Google/Chrome/User Data")
    options.add_argument(f'--profile-directory={profile}')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 2,
        "profile.default_content_setting_values.notifications": 2
    })
    driver = webdriver.Chrome(options=options)
    driver.get("https://google.com")
    return driver


#  def __init__(self, id, code, entry_time, exit_time, driver):
main_driver = open_chrome("Profile 8")
new_class = Schedule(0,"TESTING",0,0, main_driver)
new_class.exit_meeting_input()
